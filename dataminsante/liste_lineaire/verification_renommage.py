
# -*- coding: utf-8 -*-
# dataminsante/liste_lineaire/verification_renommage.py

from typing import Optional, Iterable, List
import re
import logging
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from dataminsante.database.database_pyramide import (
    code_provinces_deux_lettres,
    code_provinces_trois_lettres,
)

logger = logging.getLogger(__name__)

# =====================================================
# Codes provinces
# =====================================================
CODES_PROVINCES = {
    2: set(code_provinces_deux_lettres.keys() if isinstance(code_provinces_deux_lettres, dict) else code_provinces_deux_lettres),
    3: set(code_provinces_trois_lettres.keys() if isinstance(code_provinces_trois_lettres, dict) else code_provinces_trois_lettres),
}

# =====================================================
# Regex fichiers Resume et SOP
# =====================================================
FILE_PATTERNS = {
    "resume": {
        "brut": re.compile(r"^(?P<code>[a-z]{2})_(?P<zone>[A-Za-zéèêÉÈÊ\-]+)_LL_(?P<maladie>[A-Za-z]+)(?:_SE\d+)?_(?P<date>\d{8})\.xlsx$"),
        "fusion": re.compile(r"^(?P<code>[a-z]{2})_LL_(?P<maladie>[A-Za-z]+)(?:_SE\d+)?_(?P<date>\d{8})\.xlsx$"),
        "pattern_feuille": re.compile(r"^LL_[A-Za-z]+$"),
        "long_code": 2,
    },
    "sop": {
        "LL": r"^LLCholera_(DPS|Labo)_[A-Z]{3}(?:_ZS_[A-Za-z0-9]+)?_SE\d{2}_\d{8}\.xlsx$",
        "Contacts": r"^BDContactsCholera_(DPS|ZS)_[A-Z]{3}(?:_ZS_[A-Za-z0-9]+)?_SE\d{2}_\d{8}\.xlsx$",
        "Labo": r"^BDLaboCholera_Labo_[A-Za-z0-9]+_[A-Z]{3}_SE\d{2}_\d{8}\.xlsx$",
        "Vaccin": r"^BD(Vaccin|PCIVaccin)Cholera_(DPS|ZS)_[A-Z]{3}(?:_ZS_[A-Za-z0-9]+)?_SE\d{2}_\d{8}\.xlsx$",
        "PCI": r"^BDPCI(PPL|Score|Scorecard)?Cholera_(DPS|ZS)_[A-Z]{3}(?:_ZS_[A-Za-z0-9]+)?_SE\d{2}_\d{8}\.xlsx$",
        "RechercheActive": r"^BDRACholera_(DPS|ZS)_[A-Z]{3}(?:_ZS_[A-Za-z0-9]+)?_SE\d{2}_\d{8}\.xlsx$",
        "Journalier": r"^BDJournalierCholera_(DPS|ZS)_[A-Z]{3}_SE\d{2}_\d{8}\.xlsx$",
        "pattern_feuille": None,
        "long_code": 3,
        "PATTERN_FILE_DPS": re.compile(r"^(?P<prefix>[A-Za-z]+)_DPS_(?P<prov>[A-Z]{3})(?:_ZS_(?P<zs>[A-Za-z0-9_\-]+))?_SE(?P<se>\d{2})_(?P<date>\d{8})(?:_compiled)?\.xlsx$"),
        "PATTERN_FILE_LAB": re.compile(r"^(?P<prefix>[A-Za-z]+)_Labo_(?P<lab>[A-Za-z0-9\-]+)_(?P<prov>[A-Z]{3})_SE(?P<se>\d{2})_(?P<date>\d{8})(?:_compiled)?\.xlsx$"),
    }
}

# =====================================================
# Détection ligne d'entête
# =====================================================
def detect_header_row(
    xlsx_path: Path,
    sheet: Optional[str] = None,
    max_search_rows: int = 10,
    known_variants: Optional[Iterable[str]] = None
) -> Optional[int]:
    """Heuristique pour détecter la ligne d'entête (0-based)."""
    try:
        wb = load_workbook(xlsx_path, read_only=False, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        if ws.merged_cells.ranges:
            wb.close()
            return None

        candidates: List[int] = []
        variants_lower = {v.strip().lower() for v in (known_variants or [])}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search_rows, values_only=True)):
            vals = [str(x).strip() for x in row if x is not None and str(x).strip() != ""]
            if len(vals) >= 3:
                if not variants_lower:
                    candidates.append(i)
                else:
                    hits = sum(1 for v in vals if v.lower() in variants_lower)
                    if hits >= 1:
                        candidates.append(i)
        return min(candidates) if candidates else None
    finally:
        wb.close()

# =====================================================
# Fonction centrale
# =====================================================
def verifier_excel_recursive(
    dossier: str,
    nomenclature: str = "sop",
    mode: str = "tous",
    nom_feuille: str = None,
    afficher: bool = False,
    detecter_header: bool = False
) -> dict:
    dossier_path = Path(dossier)
    rapport = []

    if nomenclature not in FILE_PATTERNS:
        raise ValueError("nomenclature doit être 'sop' ou 'resume'")

    cfg = FILE_PATTERNS[nomenclature]

    if nomenclature == "resume":
        patterns_dict = {
            k: v for k, v in cfg.items()
            if k not in ["pattern_feuille", "long_code", "PATTERN_FILE_DPS", "PATTERN_FILE_LAB"]
        }
        pattern_feuille = cfg.get("pattern_feuille")
        long_code = cfg.get("long_code")
    else:
        patterns_dict = {k: v for k, v in cfg.items() if not k.startswith("PATTERN_FILE")}
        pattern_feuille = None
        long_code = cfg.get("long_code", None)

    # Déterminer les patterns à tester
    if mode == "tous":
        patterns = list(patterns_dict.values())
    elif mode in patterns_dict:
        patterns = [patterns_dict[mode]]
    else:
        raise ValueError(f"Mode '{mode}' invalide pour nomenclature {nomenclature}")

    for fichier in dossier_path.rglob("*.xlsx"):
        erreurs = []
        nom_fichier = fichier.name
        chemin_complet = str(fichier)
        match = None
        meta = {}
        header_row = None

        # Vérification du nom de fichier
        for p in patterns:
            if isinstance(p, re.Pattern):
                m = p.match(nom_fichier)
                if m:
                    match = m
                    break
            elif isinstance(p, str):
                if re.match(p, nom_fichier):
                    match = True
                    break

        fichier_valide = bool(match)
        if not fichier_valide:
            erreurs.append(f"Nom de fichier non conforme ({nomenclature}, mode={mode})")
        else:
            # Validation spécifique Resume
            if nomenclature == "resume" and isinstance(match, re.Match):
                meta = match.groupdict()
                code = meta.get("code")
                date = meta.get("date")
                if code and code not in CODES_PROVINCES.get(long_code, []):
                    erreurs.append(f"Code province invalide ({code})")
                if date and not re.match(r"^\d{8}$", date):
                    erreurs.append(f"Date invalide ({date})")
                fichier_valide = len(erreurs) == 0
            # Validation spécifique SOP
            elif nomenclature == "sop":
                for key in ["PATTERN_FILE_DPS", "PATTERN_FILE_LAB"]:
                    pattern = cfg.get(key)
                    if pattern:
                        m = pattern.match(nom_fichier)
                        if m:
                            meta = m.groupdict()
                            prov = meta.get("prov")
                            if prov and prov not in CODES_PROVINCES.get(long_code, []):
                                erreurs.append(f"Code province invalide ({prov})")

        # Vérification des feuilles
        try:
            wb = load_workbook(fichier, read_only=True)
            feuilles = wb.sheetnames
            if nom_feuille:
                feuille_valide = nom_feuille in feuilles
                if not feuille_valide:
                    erreurs.append(f"Nom de feuille attendu absent ({nom_feuille})")
            else:
                if nomenclature == "resume" and pattern_feuille:
                    feuilles_invalides = [s for s in feuilles if not pattern_feuille.match(s)]
                    feuille_valide = len(feuilles_invalides) == 0
                    if not feuille_valide:
                        erreurs.append(f"Feuilles invalides : {', '.join(feuilles_invalides)}")
                else:
                    feuille_valide = all(s.startswith(("LL", "BD")) for s in feuilles)
                    if not feuille_valide:
                        erreurs.append("Au moins une feuille ne correspond pas au format attendu")

            # Détection ligne header optionnelle
            if detecter_header:
                sheet_to_check = nom_feuille if nom_feuille else feuilles[0]
                header_row = detect_header_row(fichier, sheet=sheet_to_check)

        except Exception as e:
            erreurs.append(f"Erreur ouverture fichier Excel : {e}")
            feuille_valide = False
        finally:
            wb.close()

        rapport.append({
            "fichier": nom_fichier,
            "chemin_complet": chemin_complet,
            "fichier_valide": fichier_valide,
            "feuille_valide": feuille_valide,
            "erreurs": erreurs,
            "header_row": header_row,
            **meta
        })

        if afficher:
            logger.warning(rapport[-1])

    fichiers_valides = [r["chemin_complet"] for r in rapport if r["fichier_valide"]]
    fichiers_invalides = [r["chemin_complet"] for r in rapport if not r["fichier_valide"]]

    return {
        "total_fichiers": len(rapport),
        "fichiers_valides": fichiers_valides,
        "fichiers_invalides": fichiers_invalides,
        "total_feuilles_valides": sum(1 for r in rapport if r["feuille_valide"]),
        "total_feuilles_invalides": sum(1 for r in rapport if not r["feuille_valide"]),
        "details": rapport
    }

# =====================================================
# Création DataFrame résumé
# =====================================================
def creer_df_resume(resume: dict) -> pd.DataFrame:
    df_details = pd.DataFrame(resume["details"])
    if "erreurs" not in df_details.columns:
        df_details["erreurs"] = [[] for _ in range(len(df_details))]
    df_details["erreurs"] = df_details["erreurs"].apply(lambda x: "; ".join(x) if isinstance(x,list) else ("" if pd.isna(x) else str(x)))
    df_details["global_valide"] = df_details["fichier_valide"] & df_details["feuille_valide"]

    df_resume = pd.DataFrame({
        "Total fichiers":[len(df_details)],
        "Fichiers globalement valides":[df_details["global_valide"].sum()],
        "Fichiers invalides":[(~df_details["global_valide"]).sum()],
        "Feuilles valides":[df_details["feuille_valide"].sum()],
        "Feuilles invalides":[(~df_details["feuille_valide"]).sum()]
    })
    return df_resume, df_details
